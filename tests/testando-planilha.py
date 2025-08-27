from openpyxl import load_workbook

abrir = load_workbook(r"C:\Users\joao.beserra\Downloads\Power Automate - Controle de energias elétricas - 20251.xlsx")
planilha = abrir["ATUAL - POWER AUTOMATE"]

# dict para armazenar o primeiro lugar em que o ID apareceu 
chaves_ids = {}

# list para registrar as duplicatas
duplicatas = {}

# começando da linha 2 até ao maximo de linhas/ ultima linha
for linha in range(2, planilha.max_row + 1):
    # celula id é igual ao conteudo da celula da coluna 3 em casa linha
    celula_id = planilha.cell(row=linha, column=3).value
    # transformo o id em string ignorando caso nao seja none
    id_normalizado = str(celula_id).strip() if celula_id is not None else None

    # se n tiver id_normalizado ele nao insere
    if not id_normalizado:
        continue

    # parece estranho mas intancia-se o objeto ids_ocorrencias fora, logo quando houver id_normalizado duplicado, ele entra nesse objeto
    if id_normalizado in chaves_ids:
        if id_normalizado not in duplicatas:
            duplicatas[id_normalizado] = [chaves_ids[id_normalizado]]  # registra a primeira vez
        duplicatas[id_normalizado].append(linha)  # adiciona nova repetição
    else:
        chaves_ids[id_normalizado] = linha

# Print final com todos os IDs repetidos
print("\n🔁 IDS DUPLICADOS DETECTADOS:")
for id_duplicado, linhas in duplicatas.items():
    print(f"ID: {id_duplicado} ➜ Linhas: {', '.join(map(str, linhas))}")
